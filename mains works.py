import random,ast,csv,time,os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
tmerytire=time.time()
def counter(pp,e,TIME):
        kmc={}    
        for x in pp:
            km={'00': [], '01': [], '02': [], '03': [], '04': [], '05': [], '06': [], '07': [], '08': [], '09': [], '10': [], '11': [], '12': [], '13': [], '14': [], '15': [], '16': [], '17': [], '18': [], '19': [], '20': [], '21': [], '22': [], '23': [], '24': [], '25': [], '26': [], '27': [], '28': [], '29': [], '30': [], '31': [], '32': [], '33': [], '34': [], '35': [], '36': [], '37': [], '38': [], '39': [], '40': [], '41': [], '42': [], '43': [], '44': [], '45': [], '46': [], '47': [], '48': [], '49': []}
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                km[str(yy)+str(zz)]=km[str(yy)+str(zz)]+[y]
            kmc[x]=km
        kmn={}
        for y in kmc:
            filtered_data = {k: v for k, v in kmc[y].items() if len(v) >1}
            if filtered_data!={}:
                kmn[y]=filtered_data
        if abs(time.time()-TIME)>7:
            kmn={}
        return kmn
def find_words_with_substring(words_list, substring):
    result = [word for word in words_list if substring in word]
    return result
def generator_table(a,b,c,d,TIME):
    e={}
    for x in a:
        e[x]=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
    j=[]
    for x in e:
        if x in d:
            for y in d[x]:
                for z in d[x][y]:
                    if e[x][int(z[0])][int(z[1])]=='':
                        e[x][int(z[0])][int(z[1])]=y
        if x in c:
              for y in list(c[x].keys()):
                  if y not in j:
                      j.append(y)
        if x in b:
            for y in list(b[x].keys()):
                if y not in j:
                    j.append(y)
    main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
    h={}
    for x in j:
        h[x]=main
    j={}
    j=h
    v=[]
    for x in c:
        for z in c[x].keys():
            if z not in v:
                v.append(z)
    for x in j:
        main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
        if x in v:
            for y in c:
                if x in c[y]:
                    k=int(int(c[y][x])/2)
                    l=[]
                    for xx in range(len(e[y])):
                        for yy in range(len(e[y][xx])):
                            if e[y][xx][yy]=="":
                                l.append(str(xx)+str(yy))
                    l = ['01', '02', '03', '04', '06', '07', '08', '09', '11', '12', '13', '14', '16', '17', '18', '19', '21', '22', '23', '24', '26', '27', '28', '29', '31', '32', '33', '34', '36', '37', '38', '39', '41', '42', '43', '44', '46', '47', '48', '49']
                    m = [[l[i], l[i+1]] for i in range(len(l)-1) if int(l[i+1]) - int(l[i]) == 1]
                    
                    while k>0:
                        if len(m)>0:
                            s=random.choice(m)
                            if s[0] in main and s[1] in main:
                                if e[y][int(s[0][0])][int(s[0][1])]=="" and e[y][int(s[1][0])][int(s[1][1])]=="":
                                    if x not in e[y][int(s[0][0])]:
                                        e[y][int(s[0][0])][int(s[0][1])]=x
                                        e[y][int(s[1][0])][int(s[1][1])]=x
                                        main.remove(s[0])
                                        main.remove(s[1])
                                        k=k-1
                                    else:
                                        if s in m:
                                            m.remove(s)
                                else:
                                    if s in m:
                                        m.remove(s)
                            else:
                                if s in m:
                                    m.remove(s)
                        else:
                            k=0
    v=[]
    for x in b:
        for z in b[x].keys():
            if z not in v:
                v.append(z)
    g={}
    vv=[]
    for x in c:
        for z in c[x].keys():
            if z not in vv:
                vv.append(z)
    main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
    for x in v:
        if x in vv:
            if x in j:
                g[x]=j[x]
            else:
                g[x]=main
        else:
            g[x]=main
    for x in g:
        for y in b:
            if x in b[y].keys():
                l=[]
                for xx in range(len(e[y])):
                    for yy in range(len(e[y][xx])):
                        if e[y][xx][yy]=="":
                            l.append(str(xx)+str(yy))
                n=int(b[y][x])
                while n>0:
                    if len(l)>0:
                        ll=random.choice(l)
                        if ll in g[x]:
                            if e[y][int(ll[0])][int(ll[1])] == "":
                                if x not in e[y][int(ll[0])]:
                                    e[y][int(ll[0])][int(ll[1])]=x
                                    g[x].remove(ll)
                                    n=n-1
                                else:
                                    l.remove(ll)
                            else:
                                l.remove(ll)
                        else:
                            l.remove(ll)
                    else:
                        n=0
    for x in e:
        l=a[x]
        u={}
        for y in l:
            u[y]=int(l[y])
        for y in u:
            r=u[y]
            w=[]
            for xx in range(len(e[x])):
                for yy in range(len(e[x][xx])):
                    if e[x][xx][yy]=="":
                        w.append(str(xx)+str(yy))
            while r>0:
                if len(w)>0:
                    qw=random.choice(w)
                    if e[x][int(qw[0])].count(y)<3 and e[x][int(qw[0])].count(y+" PR")<4:
                        count = 0
                        for subject in e[x][int(qw[0])]:
                            if y in subject:
                                count += 1
                        if count<4:
                            if e[x][int(qw[0])][int(qw[1])]=="":
                                e[x][int(qw[0])][int(qw[1])]=y
                                r=r-1
                            else:
                                w.remove(qw)
                        else:
                            w.remove(qw)
                    else:
                        w.remove(qw)
                else:
                    r=0
    f={}
    for x in d:
        l=d[x]
        m={}
        for y in l:
            m[y]=len(l[y])
        f[x]=m
            
    g={}
    for x in e:   
        d4= a[x] |b[x] | c[x] | f[x]
        g[x]=d4
    f={}
    for x in e:
        data=e[x]
        i=g[x]
        flattened_data = [item for sublist in data for item in sublist]
        element_counts = Counter(flattened_data)
        h = dict(element_counts)
        j={}
        for y in i:
            if y in h:
                if int(int(i[y])-int(h[y]))!=0:
                    j[y]=int(i[y])-int(h[y])
            else:
                j[y]=i[y]
        for y in h:
            if y not in i:
                j[y]=0-h[y]
        if j!={}:
            f[x]=j
    if f!={}:
        for x in f:
            if len(f[x])==2 and list(f[x].keys())==['ENG',''] and sum(list(f[x].values()))==0:
                t=f[x]["ENG"]
                while t>0:
                    m=""
                    for y in range(len(e[x])):
                        for z in range(len(e[x][y])):
                            if e[x][y][z]=="":
                                m=str(y)+str(z)
                    k=list(a[x].keys())
                    l={}
                    for y in k:
                        l[y]=e[x][int(m[0])].count(y)
                    l=min(l)
                    k={}
                    for y in range(5):
                        if y!=int(m[0]):
                            k[y]=e[x][y].count(l)
                    k=max(k)
                    cl=[]
                    for y in range(len(e[x][k])):
                        if e[x][k][y]==l:
                            cl.append(str(k)+str(y))
                    if len(cl)!=0:
                        cl=random.choice(cl)
                        e[x][int(m[0])][int(m[1])]=l
                        e[x][int(cl[0])][int(cl[1])]="ENG"
                        t=t-1
                    else:
                        t=0
    f={}
    for x in d:
        l=d[x]
        m={}
        for y in l:
            m[y]=len(l[y])
        f[x]=m
    g={}
    for x in e:   
        d4= a[x] |b[x] | c[x] | f[x]
        g[x]=d4
    t={}
    for x in g:
        r={}
        for y in g[x]:
            r[y]=int(g[x][y])
        t[x]=r
    g=t.copy()    
    for x in g:
        elements = list(g[x].keys())
        data = e[x]
        flattened_data = [item for sublist in data for item in sublist]
        counts = Counter(flattened_data)
        filtered_counts = {element: counts[element] for element in elements} 
        if filtered_counts!=g[x]:
            js={}
            for y in filtered_counts:
                if filtered_counts[y]!=g[x][y]:
                    js[y]=g[x][y]-filtered_counts[y]
            for y in g[x]:
                if filtered_counts[y]!=g[x][y]:
                    if y not in js:
                        js[y]=g[x][y]-filtered_counts[y]
            count=0
            for y in range(len(e[x])):
                for z in range(len(e[x][y])):
                    if e[x][y][z]=="":
                        count-=1
            js['']=count
            if sum(js.values())==0:
                positive_points = {k: v for k, v in js.items() if v > 0}
                negative_points = {k: v for k, v in js.items() if v < 0}
                nested_list = []
                for pos_subj, pos_points in positive_points.items():
                    for neg_subj, neg_points in negative_points.items():
                        nested_list.append([pos_subj, neg_subj])
                for y in nested_list:
                    l=""
                    for yy in range(len(e[x])):
                        for zz in range(len(e[x][yy])):
                            if e[x][yy][zz]==y[1]:
                                l=str(yy)+str(zz)
                    lh=l
                    kh=list(a[x].keys())
                    df={}
                    for yy in kh:
                        for zz in range(len(e[x])):
                            if zz!=int(lh[0]):
                                if e[x][zz].count(yy)>=3:
                                    df[yy]=zz
                                    break
                    if len(df)!=0:
                        dk=random.choice(list(df))
                        df=[dk,df[dk]]
                        e[x][int(lh[0])][int(lh[1])]=dk
                        mh=[]
                        for cc in range(len(e[x][df[1]])):
                            if e[x][df[1]][cc]==df[0]:
                                mh.append(cc)
                        mh=random.choice(mh)
                        e[x][df[1]][mh]=y[0]
    if abs(time.time()-TIME)>7:
        e={}
    return e
    for x in e:
        for y in x:
            if len(y)>2:
                cout=1
    return cout
def names_0(b,e,a,c,TIME):
    bb = []
    for classes in b.values():
        for subject in classes.keys():
            if subject not in bb:
                bb.append(subject)
    
    cc=[]
    for classes in c.values():
        for subject in classes.keys():
            if subject not in cc:
                cc.append(subject)
    bc=[]
    for x in cc:
        if x not in bc and x in bb:
            bc.append(x)
    dic={}
    for y in bc:
        ds={}
        for x in e:
            for yy in range(len(e[x])):
                for zz in range(len(e[x][yy])):
                    if e[x][yy][zz]==y:
                        if str(yy)+str(zz) not in ds:
                            ds[str(yy)+str(zz)]=[x]
                        else:
                            ds[str(yy)+str(zz)]=ds[str(yy)+str(zz)]+[x]
        dic[y]=ds
    for x in dic:
        main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
        tr=list(dic[x].keys())
        tp=[]
        for y in main:
            if y not in tr and y not in tp:
                tp.append(y)
        dpd={}
        for y in dic[x]:
            if len(dic[x][y])>1:
                dpd[y]=dic[x][y][1]
        
        if len(dpd)!=0:
            dsp=list(dpd.keys())
            dsz=list(dpd.values())
            for y in range(len(dsz)):
                poi=list(a[dsz[y]].keys())
                lo=[]
                for yy in range(len(e[dsz[y]])):
                    for zz in range(len(e[dsz[y]][yy])):
                        if e[dsz[y]][yy][zz] in poi:
                            lo.append(str(yy)+str(zz))
                lp=[]
                for yy in lo:
                    if yy in tp:
                        lp.append(yy)
                if len(lp)!=0:
                    li=random.choice(lp)
                    tp.remove(li)
                    cp=dsp[y]
                    e[dsz[y]][int(cp[0])][int(cp[1])],e[dsz[y]][int(li[0])][int(li[1])]=e[dsz[y]][int(li[0])][int(li[1])],e[dsz[y]][int(cp[0])][int(cp[1])]
    if abs(time.time()-TIME)>7:
        e={}   
    return e             
def names_1(pp,a,e,ii,TIME):
    lencho=[]    
    while True:
        kmc={}    
        for x in pp:
            km={'00': [], '01': [], '02': [], '03': [], '04': [], '05': [], '06': [], '07': [], '08': [], '09': [], '10': [], '11': [], '12': [], '13': [], '14': [], '15': [], '16': [], '17': [], '18': [], '19': [], '20': [], '21': [], '22': [], '23': [], '24': [], '25': [], '26': [], '27': [], '28': [], '29': [], '30': [], '31': [], '32': [], '33': [], '34': [], '35': [], '36': [], '37': [], '38': [], '39': [], '40': [], '41': [], '42': [], '43': [], '44': [], '45': [], '46': [], '47': [], '48': [], '49': []}
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                km[str(yy)+str(zz)]=km[str(yy)+str(zz)]+[y]
            kmc[x]=km
        kmn={}
        for y in kmc:
            filtered_data = {k: v for k, v in kmc[y].items() if len(v) >1}
            if filtered_data!={}:
                kmn[y]=filtered_data  
        if len(kmn)==0:
            break
        lencho.append(len(kmn))
        knz={}
        for y1 in kmn:
            dsa={}
            for y2 in kmn[y1]:
                dsa[y2]=random.choices(kmn[y1][y2], k=2)
            knz[y1]=dsa
        kmn=knz.copy()
        for x in kmn:    
            for y in kmn[x]:
                hgf=""
                for r in kmn[x][y]:
                    if e[r][int(y[0])][int(y[1])] in a[r].keys():
                           hgf=r
                           break
                if len(hgf)>1:
                    qwa=""
                    for t in ii[hgf]:
                        if ii[hgf][t]==x:
                            qwa=t
                    waz=list(a[hgf].keys())
                    if qwa in waz:
                        waz.remove(qwa)
                        l=[]
                        for te in range(len(e[hgf])):
                            for tes in range(len(e[hgf][te])):
                                if e[hgf][te][tes] in waz:
                                    l.append(str(te)+str(tes))
                        kkz=[]
                        for xx in pp[x]:
                            for yy in pp[x][xx]:
                                for z1 in range(len(e[xx])):
                                    for z2 in range(len(e[xx][z1])):
                                        if e[xx][z1][z2]==yy:
                                            if str(z1)+str(z2) not in kkz:
                                                kkz.append(str(z1)+str(z2))
                        kkf=[]
                        for sa in l:
                            if sa not in kkz:
                                kkf.append(sa)
                        while len(kkf)>0:
                            kkc=random.choice(kkf)
                            eqa=e[hgf][int(kkc[0])][int(kkc[1])]
                            if eqa in ii[hgf]:
                                kkz=[]
                                for xx in pp[ii[hgf][eqa]]:
                                    for yy in pp[ii[hgf][eqa]][xx]:
                                        if yy[-3:]!=" PR":
                                            for z1 in range(len(e[xx])):
                                                for z2 in range(len(e[xx][z1])):
                                                    if e[xx][z1][z2]==yy:
                                                        if str(z1)+str(z2) not in kkz:
                                                            kkz.append(str(z1)+str(z2))
                                if y not in kkz:
                                    if e[hgf][int(kkc[0])][int(kkc[1])][-3:]!=" PR" and e[hgf][int(y[0])][int(y[1])]!=" PR":
                                        e[hgf][int(kkc[0])][int(kkc[1])],e[hgf][int(y[0])][int(y[1])]=e[hgf][int(y[0])][int(y[1])],e[hgf][int(kkc[0])][int(kkc[1])]
                                        break
                            kkf.remove(kkc)
        if len(lencho)>7:
            if lencho[-1]==lencho[-2]==lencho[-3]==lencho[-4]==lencho[-5]:
                break
        if len(lencho)>7:
            break
    if abs(time.time()-TIME)>7:
        e={}        
    return e
def names_2(pp,e,a,TIME):
    for qwsas in range(2):
        kmc={}    
        for x in pp:
            km={'00': [], '01': [], '02': [], '03': [], '04': [], '05': [], '06': [], '07': [], '08': [], '09': [], '10': [], '11': [], '12': [], '13': [], '14': [], '15': [], '16': [], '17': [], '18': [], '19': [], '20': [], '21': [], '22': [], '23': [], '24': [], '25': [], '26': [], '27': [], '28': [], '29': [], '30': [], '31': [], '32': [], '33': [], '34': [], '35': [], '36': [], '37': [], '38': [], '39': [], '40': [], '41': [], '42': [], '43': [], '44': [], '45': [], '46': [], '47': [], '48': [], '49': []}
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                km[str(yy)+str(zz)]=km[str(yy)+str(zz)]+[y]
            kmc[x]=km
        kmn={}
        for y in kmc:
            filtered_data = {k: v for k, v in kmc[y].items() if len(v) >1}
            if filtered_data!={}:
                kmn[y]=filtered_data
        if len(kmn)==0:
            break
        elif len(kmn)!=0:
            kmh={}
            kmb={}
            for x in kmn:  
                for y in kmn[x]:
                    hgf=""
                    for r in kmn[x][y]:
                        if e[r][int(y[0])][int(y[1])] in a[r].keys():
                               hgf=r
                               break
                    if len(hgf)==0:
                        pfp={}
                        pfb={}
                        for yy in kmn[x]:
                            pfp[yy]=kmn[x][yy]
                            pfb[yy]=kmn[x][yy][1]
                        kmh[x]=pfp
                        kmb[x]=pfb
            l=[]
            for x in kmh:
                for y in kmh[x]:
                    for z in kmh[x][y]:
                        if e[z][int(y[0])][int(y[1])] not in l:
                           l.append(e[z][int(y[0])][int(y[1])])
            if "COMP.PR.1" in l or "COMP.PR.2" in l:
                for x in kmb:
                    kkr=list(kmb[x].values())
                    kkc=[]
                    for y in kkr:
                        if y not in kkc:
                            kkc.append(y)
                    for y in kkc:
                        trs=[]
                        
                        for yz in range(len(e[y])):
                            for yx in range(len(e[y][yz])):
                                if  "COMP.PR.1" in e[y][yz][yx] or "COMP.PR.2" in e[y][yz][yx]:
                                    trs.append(str(yz)+str(yx))
                        
                        lk=[]
                        for yz in range(len(e[y])):
                            for yx in range(len(e[y][yz])):
                                if e[y][yz][yx] in list(a[y].keys()):
                                    lk.append(str(yz)+str(yx))
                        lmj=[]
                        for g in pp[x]:
                            for gg in pp[x][g]:
                                for yz in range(len(e[g])):
                                    for zy in range(len(e[g][yz])):
                                        if e[g][yz][zy]==gg and str(yz)+str(zy) not in lmj:
                                            lmj.append(str(yz)+str(zy))
                        lmm=[]
                        for yz in lk:
                            if yz not in lmj:
                                lmm.append(yz)
                        lkt=[]
                        for yz in e:
                            for zy in range(len(e[yz])):
                                for yt in range(len(e[yz][zy])):
                                    if "COMP.PR.1" in e[yz][zy][yt] or "COMP.PR.2" in e[yz][zy][yt]:
                                        lkt.append(str(zy)+str(yt))
                                        if str(zy)+str(yt) in lkt and str(zy)+str(yt) in lmm:
                                            lmm.remove(str(zy)+str(yt))
                        numbers = lmm.copy()
                        numbers = [int(num) for num in numbers]
                        pair_with_diff_1 = None
                        for i in range(len(numbers)):
                            for j in range(i + 1, len(numbers)):
                                if abs(numbers[i] - numbers[j]) == 1:
                                    pair_with_diff_1 = (numbers[i], numbers[j])
                                    break
                            if pair_with_diff_1:
                                break
                        if pair_with_diff_1:
                            pk=list(pair_with_diff_1)
                            phg=[]
                            for pm in pk:
                                pa=str(pm)
                                if len(pa)==1:
                                    pa="0"+pa
                                phg.append(pa)
                            if len(trs)==2 and len(phg)==2:
                                e[y][int(trs[0][0])][int(trs[0][1])],e[y][int(phg[0][0])][int(phg[0][1])]=e[y][int(phg[0][0])][int(phg[0][1])],e[y][int(trs[0][0])][int(trs[0][1])]
                                e[y][int(trs[1][0])][int(trs[1][1])],e[y][int(phg[1][0])][int(phg[1][1])]=e[y][int(phg[1][0])][int(phg[1][1])],e[y][int(trs[1][0])][int(trs[1][1])]
                                break
    if abs(time.time()-TIME)>7:
        e={}                              
    return e
def names_3(pp,e,a,ii,jj,TIME):
    for qwertx in range(2):
        kmc={}    
        for x in pp:
            km={'00': [], '01': [], '02': [], '03': [], '04': [], '05': [], '06': [], '07': [], '08': [], '09': [], '10': [], '11': [], '12': [], '13': [], '14': [], '15': [], '16': [], '17': [], '18': [], '19': [], '20': [], '21': [], '22': [], '23': [], '24': [], '25': [], '26': [], '27': [], '28': [], '29': [], '30': [], '31': [], '32': [], '33': [], '34': [], '35': [], '36': [], '37': [], '38': [], '39': [], '40': [], '41': [], '42': [], '43': [], '44': [], '45': [], '46': [], '47': [], '48': [], '49': []}
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                km[str(yy)+str(zz)]=km[str(yy)+str(zz)]+[y]
            kmc[x]=km
        kmn={}
        for y in kmc:
            filtered_data = {k: v for k, v in kmc[y].items() if len(v) >1}
            if filtered_data!={}:
                kmn[y]=filtered_data 
        if len(kmn)==0:
            break
        knz={}
        for y1 in kmn:
            dsa={}
            for y2 in kmn[y1]:
                dsa[y2]=random.choices(kmn[y1][y2], k=2)
            knz[y1]=dsa
        kmn=knz.copy()
        for x in kmn:    
            for y in kmn[x]:
                hgf=""
                for r in kmn[x][y]:
                    if e[r][int(y[0])][int(y[1])] in a[r].keys():
                           hgf=r
                           break
                if len(hgf)>1:
                    bn=e[hgf][int(y[0])][int(y[1])]
                    lo=[]
                    for yy in pp[x]:
                        for zz in pp[x][yy]:
                            for xx in range(len(e[yy])):
                                for xy in range(len(e[yy][xx])):
                                    if e[yy][xx][xy]==zz:
                                        if str(xx)+str(xy) not in lo:
                                            lo.append(str(xx)+str(xy))
                    lo=sorted(lo)
                    poj=list(ii[hgf].keys())
                    pok={}
                    for xx in poj:
                        ls=[]
                        for yy in range(len(e[hgf])):
                            for zz in range(len(e[hgf][yy])):
                                if e[hgf][yy][zz]==xx and str(yy)+str(zz) not in lo:
                                    ls.append(str(yy)+str(zz))
                        if len(ls)>0:
                            pok[xx]=ls
                    for xx in pok:
                        roi=[]
                        for yy in jj[ii[hgf][xx]]:
                            for zz in jj[ii[hgf][xx]][yy]:
                                for xy in range(len(e[yy])):
                                    for yx in range(len(e[yy][xy])):
                                        if e[yy][xy][yx]==zz and str(xy)+str(yx) not in roi:
                                            roi.append(str(xy)+str(yx))
                        if bn not in roi and len(pok[xx])>0:
                            poy=random.choice(pok[xx])
                            e[hgf][int(poy[0])][int(poy[1])],e[hgf][int(y[0])][int(y[1])]=e[hgf][int(y[0])][int(y[1])],e[hgf][int(poy[0])][int(poy[1])]            
                            break
    if abs(time.time()-TIME)>7:
        e={}       
    return e
def names_4(pp,e,b,a,TIME):
    kmn=counter(pp,e,TIME)
    if len(kmn)!=0:
        for x in kmn:
            for y in kmn[x]:
                l=[]
                for z in kmn[x][y]:
                    if e[z][int(y[0])][int(y[1])] in b[z]:
                            l.append(e[z][int(y[0])][int(y[1])])
                            l.append(y)
                            l.append(z)
                            break
                if len(l)!=0:
                    lp=[]
                    for xt in e:
                        for yt in range(len(e[xt])):
                            for zt in range(len(e[xt][yt])):
                                if e[xt][yt][zt]==l[0]:
                                    lp.append(str(yt)+str(zt))
                    lj=[]
                    for xt in range(len(e[l[2]])):
                        for yt in range(len(e[l[2]][xt])):
                            if e[l[2]][xt][yt] in a[l[2]]:
                                lj.append(str(xt)+str(yt))
                    ln=[]
                    for xt in lj:
                        if xt not in lp:
                            if xt not in ln:
                                ln.append(xt)
                    if len(ln)!=0:
                        ls=random.choice(ln)
                        phg=l[1]
                        phh=l[2]
                        e[phh][int(ls[0])][int(ls[1])],e[phh][int(phg[0])][int(phg[1])]=e[phh][int(phg[0])][int(phg[1])],e[phh][int(ls[0])][int(ls[1])]
    if abs(time.time()-TIME)>7:
        e={}        
    return e
def editor(a,b,c,d,e,TIME):
    count=0
    while count<1:
        f={}
        for x in d:
            l=d[x]
            m={}
            for y in l:
                m[y]=len(l[y])
            f[x]=m
        g={}
        for x in e:   
            d4= a[x] |b[x] | c[x] | f[x]
            g[x]=d4
        t={}
        for x in g:
            r={}
            for y in g[x]:
                r[y]=int(g[x][y])
            t[x]=r
        g=t.copy()
        for x in g:
            elements = list(g[x].keys())
            data = e[x]
            flattened_data = [item for sublist in data for item in sublist]
            counts = Counter(flattened_data)
            filtered_counts = {element: counts[element] for element in elements} 
            if filtered_counts!=g[x]:
                e=generator_table(a,b,c,d,TIME)
            else:
                count=1
    if abs(time.time()-TIME)>7:
        e={}        
    return e
def final(e,TIME):

    hsd=[]
    while len(counter(pp,e,TIME))>0:
        functions = [lambda e: names_0(b,e,a,c,TIME),lambda e: names_1(pp,a,e,ii,TIME),lambda e: names_2(pp,e,a,TIME),lambda e: names_3(pp,e,a,ii,jj,TIME),lambda e: names_4(pp,e,b,a,TIME)]
        random.shuffle(functions)
        for func in functions:
            e = func(e)
        hsd.append(len(counter(pp,e,TIME)))
        if len(hsd)>3:
            if hsd[-1]==hsd[-2]==hsd[-3]:
                break
        if len(hsd)>7:
            break
    if abs(time.time()-TIME)>7:
        e={}       
    return e
def count_subjects(nested_list):
    flattened_list = [subject for day in nested_list for subject in day]
    subject_count = Counter(flattened_list)
    subject_count_str = {subject: str(count) for subject, count in subject_count.items()}
    return subject_count_str
def adder(d,e,aas,bbs,ccs,b,c,a,TIME):
    dds={}
    for x in d:
        km={}
        for y in d[x]:
            km[y]=str(len(d[x][y]))
        dds[x]=km
    zz={}
    for x in aas:
        zz[x]=aas[x] |bbs[x] | ccs[x]| dds[x]
    for x in e:
        s = count_subjects(e[x])
        if s!=zz[x]:
            ds = [k for k in s if k in zz[x] and s[k] != zz[x][k]]
            ds.extend([k for k in zz[x] if k not in s])
            
            for y in ds:
                if y in b[x]:
                    khg=[]
                    for xx in range(len(e[x])):
                        for yy in range(len(e[x][xx])):
                            if e[x][xx][yy]=='':
                                khg.append(str(xx)+str(yy))
                    khg=random.choice(khg)
                    e[x][int(khg[0])][int(khg[1])]=y
                if y in c[x]:
                    khg=[]
                    for xx in range(len(e[x])):
                        for yy in range(len(e[x][xx])):
                            if e[x][xx][yy]=='':
                                khg.append(str(xx)+str(yy))
                    if len(khg)>1:
                        khg = random.sample(khg, 2)
                        if abs(int(khg[0])-int(khg[1]))==1:
                            e[x][int(khg[0][0])][int(khg[0][1])]=y
                            e[x][int(khg[1][0])][int(khg[1][1])]=y
                        else:
                            if e[x][int(khg[0][0])][int(khg[0][1])-1] in a[x]:
                                e[x][int(khg[0][0])][int(khg[0][1])]=y
                                e[x][int(khg[1][0])][int(khg[1][1])]=e[x][int(khg[0][0])][int(khg[0][1])-1]
                                e[x][int(khg[0][0])][int(khg[0][1])-1]=y
                            
                            elif int(khg[0][1])+1<10:
                                if e[x][int(khg[0][0])][int(khg[0][1])+1] in a[x]:
                                    e[x][int(khg[0][0])][int(khg[0][1])]=y
                                    e[x][int(khg[1][0])][int(khg[1][1])]=e[x][int(khg[0][0])][int(khg[0][1])+1]
                                    e[x][int(khg[0][0])][int(khg[0][1])+1]=y
                            
                            
                            elif e[x][int(khg[1][0])][int(khg[1][1])-1] in a[x]:
                                e[x][int(khg[1][0])][int(khg[1][1])]=y
                                e[x][int(khg[0][0])][int(khg[0][1])]=e[x][int(khg[1][0])][int(khg[1][1])-1]
                                e[x][int(khg[1][0])][int(khg[1][1])-1]=y
                            
                            elif int(khg[1][1])+1<10:
                                if e[x][int(khg[1][0])][int(khg[1][1])+1] in a[x]:
                                    e[x][int(khg[1][0])][int(khg[1][1])]=y
                                    e[x][int(khg[0][0])][int(khg[0][1])]=e[x][int(khg[1][0])][int(khg[1][1])+1]
                                    e[x][int(khg[1][0])][int(khg[1][1])+1]=y
    zz={}
    vc=0
    for x in aas:
        zz[x]=aas[x] |bbs[x] | ccs[x]| dds[x]
    
    for x in e:
        s = count_subjects(e[x])
        if s!=zz[x]:
            vc=1
            break
    if abs(time.time()-TIME)>7:
        e={}       
    return vc,e
def find_pairs(lst):
    lst.sort()
    pairs = []
    
    for i in range(len(lst) - 1):
        if int(lst[i+1]) - int(lst[i]) == 1:
            pairs.append([lst[i], lst[i+1]])
        pass
    return pairs
def setter(l,lm,x,op,e,a,b,c,TIME):
    du = []
    seen = []
    for number in l:
        if number in seen and number not in du:
            du.append(number)
        else:
            seen.append(number)
    if du!=[]:
        kh={}
        for t in lm:
            if t in du:
                kh[t]=lm[t]
        if x in op:
            for gg in du:
                l.remove(gg)
            dc=find_pairs(du)
            for y in dc:
                if y[0] in kh and y[1] in kh:
                    if kh[y[0]]==kh[y[1]]:
                        xc=kh[y[0]]
                        ln=[]
                        for xx in range(len(e[xc])):
                            for yy in range(len(e[xc][xx])):
                                if e[xc][xx][yy] in a[xc] or e[xc][xx][yy] in b[xc]:
                                    if str(xx)+str(yy) not in l:
                                        ln.append(str(xx)+str(yy))
                        if len(find_pairs(ln))!=0:
                            ln=random.choice(find_pairs(ln))
                            if len(ln)==2 and len(y)==2:
                                e[xc][int(ln[0][0])][int(ln[0][1])],e[xc][int(y[0][0])][int(y[0][1])]=e[xc][int(y[0][0])][int(y[0][1])],e[xc][int(ln[0][0])][int(ln[0][1])]
                                e[xc][int(ln[1][0])][int(ln[1][1])],e[xc][int(y[1][0])][int(y[1][1])]=e[xc][int(y[1][0])][int(y[1][1])],e[xc][int(ln[1][0])][int(ln[1][1])]
                                if y[0] in du and y[1] in du:
                                    du.remove(y[0])
                                    du.remove(y[1])
                                    l.append(ln[0])
                                    l.append(ln[1])
                                    kh.pop(y[0])
                                    kh.pop(y[1])
            l=sorted(l)
            kj={}
            for y in kh:
                kj[kh[y]]=y
            for y in kj:
                if e[y][int(kj[y][0])][int(kj[y][1])] in c[y]:
                    lk=[]
                    for xx in range(len(e[y])):
                        for yy in range(len(e[y][xx])):
                            if e[y][xx][yy]==x:
                                lk.append(str(xx)+str(yy))
                    if len(find_pairs(lk))!=0:
                        lk=random.choice(find_pairs(lk))
                        xc=y
                        ln=[]
                        for xx in range(len(e[xc])):
                            for yy in range(len(e[xc][xx])):
                                if e[xc][xx][yy] in a[xc] or e[xc][xx][yy] in b[xc]:
                                    if str(xx)+str(yy) not in l:
                                        ln.append(str(xx)+str(yy))
                        
                        if len(find_pairs(ln))!=0:
                            ln=random.choice(find_pairs(ln))
                            y=lk.copy()
                            if len(ln)==2 and len(y)==2:
                                e[xc][int(ln[0][0])][int(ln[0][1])],e[xc][int(y[0][0])][int(y[0][1])]=e[xc][int(y[0][0])][int(y[0][1])],e[xc][int(ln[0][0])][int(ln[0][1])]
                                e[xc][int(ln[1][0])][int(ln[1][1])],e[xc][int(y[1][0])][int(y[1][1])]=e[xc][int(y[1][0])][int(y[1][1])],e[xc][int(ln[1][0])][int(ln[1][1])]
                                if y[0] in du and y[1] in du and xc in kh:
                                    du.remove(y[0])
                                    du.remove(y[1])
                                    l.append(ln[0])
                                    l.append(ln[1])
                                    kh.pop(xc)
        for y in kh:
            xc=kh[y]
            if x in b[xc]:
                lb=[]
                for yy in range(len(e[xc])):
                    for zz in range(len(e[xc][yy])):
                        if e[xc][yy][zz] in a[xc]:
                            if str(yy)+str(zz) not in l:
                                lb.append(str(yy)+str(zz))
                if len(lb)>0:
                    lb=random.choice(lb)
                    l.append(lb)
                    e[xc][int(lb[0])][int(lb[1])],e[xc][int(y[0])][int(y[1])]=e[xc][int(y[0])][int(y[1])],e[xc][int(lb[0])][int(lb[1])]
     
    if abs(time.time()-TIME)>7:
        e={}        
    return e
def setter_2(e,c,a,b,TIME):
    tt=[]
    for x in b:
        for y in b[x]:
            if y not in tt:
                tt.append(y)
    for x in c:
        for y in c[x]:
            if y not in tt:
                tt.append(y)
    op=[]
    for re in list(c.values()):
        re=list(re.keys())
        for rs in re:
            if rs not in op:
                op.append(rs)
    for x in tt:
        lm={}
        l=[]
        for xx in e:
            for yy in range(len(e[xx])):
                for zz in range(len(e[xx][yy])):
                    if e[xx][yy][zz]==x:
                        l.append(str(yy)+str(zz))
                        lm[str(yy)+str(zz)]=xx
        
        e=setter(l,lm,x,op,e,a,b,c,TIME)
     
    if abs(time.time()-TIME)>7:
        e={}       
    return e
def setter_3(e,c,a,b,TIME):
    tt=[]
    for x in b:
        for y in b[x]:
            if y not in tt:
                tt.append(y)
    for x in c:
        for y in c[x]:
            if y not in tt:
                tt.append(y)
    op=[]
    for re in list(c.values()):
        re=list(re.keys())
        for rs in re:
            if rs not in op:
                op.append(rs)
    for x in tt:
        lm={}
        l=[]
        for xx in e:
            for yy in range(len(e[xx])):
                for zz in range(len(e[xx][yy])):
                    if e[xx][yy][zz]==x:
                        l.append(str(yy)+str(zz))
                        if str(yy)+str(zz) not in lm:
                            lm[str(yy)+str(zz)]=xx
        
        e=setter(l,lm,x,op,e,a,b,c,TIME)
     
    if abs(time.time()-TIME)>7:
        e={}   
    return e
def checks(a,b,c,e,TIME):
    if abs(time.time()-TIME)>7:
        e={}
    tms=[]       
    tt=[]
    for x in b:
        for y in b[x]:
            if y not in tt:
                tt.append(y)
    for x in c:
        for y in c[x]:
            if y not in tt:
                tt.append(y)
    op=[]
    for re in list(c.values()):
        re=list(re.keys())
        for rs in re:
            if rs not in op:
                op.append(rs)
    for x in tt:
        lm={}
        l=[]
        for xx in e:
            for yy in range(len(e[xx])):
                for zz in range(len(e[xx][yy])):
                    if e[xx][yy][zz]==x:
                        l.append(str(yy)+str(zz))
                        lm[str(yy)+str(zz)]=xx   
        du = []
        seen = []
        for number in l:
            if number in seen and number not in du:
                du.append(number)
            else:
                seen.append(number)
        if du!=[]:
            tms=tms+du
    if len(tms)<5:
        return 0
    else:
        return 1
def prt_changer(a,e):
    l=[]
    for x in e:
        for y in range(len(e[x])):
            if e[x][y][9]=="CLUBS":
                if y not in l:
                    l.append(y)
    ll=[]
    for x in e:
        for y in l:
            if e[x][y][9]=='COMP.PR.2' or e[x][y][9]=='COMP.PR.1':
                ll.append([x,y,e[x][y][9]])
    for x in ll:
        main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
        for y in e:
            for z in range(len(e[y])):
                for u in range(len(e[y][z])):
                    if x[2] in e[y][z][u]:
                        if str(z)+str(u) in main:
                            main.remove(str(z)+str(u))
        main=find_pairs(main)
        ls=[]
        for y in range(len(e[x[0]])):
            for z in range(len(e[x[0]][y])):
                if e[x[0]][y][z] in a[x[0]]:
                    ls.append(str(y)+str(z))
        ls=find_pairs(ls)
        lk=[]
        for y in ls:
            if y in main:
                lk.append(y)
        if len(lk)>0:
            lk=random.choice(lk)
            ab=lk[0]
            bc=lk[1]
            cd=x[1]
            de=x[0]
            e[de][cd][9],e[de][int(bc[0])][int(bc[1])]=e[de][int(bc[0])][int(bc[1])],e[de][cd][9]
            e[de][cd][8],e[de][int(ab[0])][int(ab[1])]=e[de][int(ab[0])][int(ab[1])],e[de][cd][8]
    return e
def convert_string_list(input_list):
    return [ast.literal_eval(item) for item in input_list]
def tuples_to_dict(tuples_list):
    return {subject: hours for subject, hours in tuples_list}
def fn(ff,e,c,b,a,TIME):
    for x in ff:
        t=count_subjects(e[x])
        D={}
        for y in ff[x]:
            D[y]=int(ff[x][y])
        ff[x]={}
        for y in D:
            ff[x][y]=D[y]
        D={}
        for y in t:
            D[y]=int(t[y])
        t={}
        for y in D:
            t[y]=D[y]
        mn={}
        for y in ff[x]:
            if y not in t:
                mn[y]=ff[x][y]
            elif ff[x][y]!=t[y]:
                mn[y]=ff[x][y]-t[y]
        
        if mn!={}:
            for y in mn:
                if y in c[x]:
                    if mn[y]>0:
                        for yy in range(len(e[x])):
                            for zz in range(len(e[x][yy])):
                                if e[x][yy][zz]=="":
                                    if zz<9:
                                        if e[x][yy][zz+1]=="":
                                            e[x][yy][zz+1]=y
                                            e[x][yy][zz]=y
                                            mn[y]=mn[y]-2
                if y in b[x]:
                    if mn[y]>0:
                        for yy in range(len(e[x])):
                            for zz in range(len(e[x][yy])):
                                if e[x][yy][zz]=="":
                                    e[x][yy][zz]=y
                                    mn[y]=mn[y]-1
            for y in mn:
                if mn[y]<0:
                    gf=mn[y]
                    gf=abs(gf)
                    lo=[]
                    for yy in range(len(e[x])):
                        for zz in range(len(e[x][yy])):
                            if e[x][yy][zz]==y:
                                lo.append(str(yy)+str(zz))
                    
                    lo = random.sample(lo, gf)
                    for z in mn:
                        if mn[z]>=gf:
                            pass
                    for z in lo:
                        e[x][int(z[0])][int(z[1])]=""
                        
            mk=list(mn)
            for y in mk:
                if y in mn:
                    if mn[y]<=0:
                        mn.pop(y)
            if mn!={}:
                for y in mn:
                    if y in c[x]:
                        khg=[]
                        for xx in range(len(e[x])):
                            for yy in range(len(e[x][xx])):
                                if e[x][xx][yy]=='':
                                    khg.append(str(xx)+str(yy))
                        if len(khg)>1:
                            khg = random.sample(khg, 2)
                            if abs(int(khg[0])-int(khg[1]))==1:
                                e[x][int(khg[0][0])][int(khg[0][1])]=y
                                e[x][int(khg[1][0])][int(khg[1][1])]=y
                            else:
                                if e[x][int(khg[0][0])][int(khg[0][1])-1] in a[x]:
                                    e[x][int(khg[0][0])][int(khg[0][1])]=y
                                    e[x][int(khg[1][0])][int(khg[1][1])]=e[x][int(khg[0][0])][int(khg[0][1])-1]
                                    e[x][int(khg[0][0])][int(khg[0][1])-1]=y
                                
                                elif int(khg[0][1])+1<10:
                                    if e[x][int(khg[0][0])][int(khg[0][1])+1] in a[x]:
                                        e[x][int(khg[0][0])][int(khg[0][1])]=y
                                        e[x][int(khg[1][0])][int(khg[1][1])]=e[x][int(khg[0][0])][int(khg[0][1])+1]
                                        e[x][int(khg[0][0])][int(khg[0][1])+1]=y
                                
                                
                                elif e[x][int(khg[1][0])][int(khg[1][1])-1] in a[x]:
                                    e[x][int(khg[1][0])][int(khg[1][1])]=y
                                    e[x][int(khg[0][0])][int(khg[0][1])]=e[x][int(khg[1][0])][int(khg[1][1])-1]
                                    e[x][int(khg[1][0])][int(khg[1][1])-1]=y
                                
                                elif int(khg[1][1])+1<10:
                                    if e[x][int(khg[1][0])][int(khg[1][1])+1] in a[x]:
                                        e[x][int(khg[1][0])][int(khg[1][1])]=y
                                        e[x][int(khg[0][0])][int(khg[0][1])]=e[x][int(khg[1][0])][int(khg[1][1])+1]
                                        e[x][int(khg[1][0])][int(khg[1][1])+1]=y
    if abs(time.time()-TIME)>7:
        e={}
    return e
def output(a,b,c,d,pp,aas,bbs,ccs,ff,TIME):
    e= final(editor(a,b,c,d,generator_table(a,b,c,d,TIME),TIME),TIME)
    
    while checks(a,b,c,e,TIME)!=0:
        kc=1
        while len(counter(pp,e,TIME))!=0 or kc==1:
            e= final(editor(a,b,c,d,generator_table(a,b,c,d,TIME),TIME),TIME)
            kc,e=adder(d,e,aas,bbs,ccs,b,c,a,TIME)
            
         
        for qw in range(2):
            functions = [lambda e: setter_2(e,c,a,b,TIME),lambda e: setter_3(e,c,a,b,TIME)]
            random.shuffle(functions)
            for func in functions:
                e = func(e)   
    e=fn(ff,e,c,b,a,TIME)        
    e= final(e,TIME)   
    while checks(a,b,c,e,TIME)!=0:
        functions = [lambda e: setter_2(e,c,a,b,TIME),lambda e: setter_3(e,c,a,b,TIME)]
        random.shuffle(functions)
        for func in functions:
            e = func(e)
    e= final(e,TIME)
     
    if abs(time.time()-TIME)>7:
        e={}
    return e
def non_movable(e,d,TIME):
    lm=[]   
    for x in e:
        for y in d[x]:
            for z in d[x][y]:
                if e[x][int(z[0])][int(z[1])]!=y:
                    lm.append("1")
    if len(lm)!=0:
        e={}
    if abs(time.time()-TIME)>7:
        e={}
    return e
def doubler(dx,e,a,TIME):
    if len(dx)!=0:
        for x in dx:
            for y in dx[x]:
                ctp=[]
                for yy in range(len(e[x])):
                    for zz in range(len(e[x][yy])):
                        if e[x][yy][zz]=="":
                            ctp.append(str(yy)+str(zz))
                if len(ctp)>=int(dx[x][y]):
                    m=int(int(dx[x][y])/2)
                    
                    while m>0 and len(ctp)>0:
                        ctt=random.sample(ctp, 2)
                        for z in ctt:
                            ctp.remove(z)
                        if e[x][int(ctt[0][0])][int(ctt[0][1])]=="" and e[x][int(ctt[1][0])][int(ctt[1][1])]=="":
                            u=ctt[0]
                            v=ctt[1]
                            if int(u[1])>0:
                                if e[x][int(u[0])][int(u[1])-1] in a[x]:
                                    e[x][int(v[0])][int(v[1])]=e[x][int(u[0])][int(u[1])-1]
                                    e[x][int(u[0])][int(u[1])-1]=y
                                    e[x][int(u[0])][int(u[1])]=y
                            elif int(u[1])<9:
                                if e[x][int(u[0])][int(u[1])+1] in a[x]:
                                    e[x][int(v[0])][int(v[1])]=e[x][int(u[0])][int(u[1])+1]
                                    e[x][int(u[0])][int(u[1])+1]=y
                                    e[x][int(u[0])][int(u[1])]=y
                            elif int(v[1])>0:
                                if e[x][int(v[0])][int(v[1])-1] in a[x]:
                                    e[x][int(u[0])][int(u[1])]=e[x][int(v[0])][int(v[1])-1]
                                    e[x][int(v[0])][int(v[1])-1]=y
                                    e[x][int(v[0])][int(v[1])]=y
                            elif int(v[1])<9:
                                if e[x][int(v[0])][int(v[1])+1] in a[x]:
                                    e[x][int(u[0])][int(u[1])]=e[x][int(v[0])][int(v[1])+1]
                                    e[x][int(v[0])][int(v[1])+1]=y
                                    e[x][int(v[0])][int(v[1])]=y
                        m=m-1
    if abs(time.time()-TIME)>7:
        e=1
    return e
def helop(e,a,b,c,TIME):
    foul=0
    while foul==0:
        ds=[]
        dm=[]
        df=open("INPUT1.csv","r",newline="")
        wr=csv.reader(df)
        for x in wr:
            while "" in x:
                x.remove("")
            if x!=[]:
                if x[0][0]!="-":
                    if len(x)==2:
                        x=[{x[0]:x[1]}]
                    elif len(x)==3:
                        x=[{x[0]:x[1]},{x[0]:x[2]}]
                    elif len(x)==4:
                        x=[{x[0]:x[1]},{x[0]:x[2]},{x[0]:x[3]}]
                    elif len(x)==5:
                        x=[{x[0]:x[1]},{x[0]:x[2]},{x[0]:x[3]},{x[0]:x[4]}]
                    else:
                        x=x[0]
                        dm.append(x)
                    ds.append(x)
        hh={}
        for x in dm:
            if x[0]=="I" or x[0]=="V" or x[0]=="X":
                n=ds.index(x)
                j=1
                DS={}
                while ds[n+j]==list(ds[n+j]):
                    m=ds[n+j]
                    for y in m:
                        DS.update(y)
                    if n+j==len(ds)-1:
                        break
                    j+=1
                if x[0]=="I" or x[0]=="V" or x[0]=="X":
                    hh[x]=DS
        ll=[]
        lg=[]
        lf=[]
        for x in hh:
            for y in hh[x]:
                if hh[x][y] not in ll:
                    ll.append(hh[x][y])
        for x in ll:
            lp=[]
            for y in hh:
                for z in hh[y]:
                    if hh[y][z]==x:
                        for yy in range(len(e[y])):
                            for zz in range(len(e[y][yy])):
                                if e[y][yy][zz]==z:
                                    lp.append(str(yy)+str(zz))
            
            items = lp.copy()
            item_counts = Counter(items)
            duplicates = {item: count for item, count in item_counts.items() if count > 1}
            if len(duplicates)!=0:
                lg.append(x)
                for t in duplicates:
                    tt={}
                    for y in hh:
                        if x in hh[y].values():
                            for z in hh[y]:
                                if hh[y][z]==x:
                                    if e[y][int(t[0])][int(t[1])]==z:
                                        if z in a[y] or z in b[y] or z in c[y]:
                                            tt[y]=z
                    lf.append(x)
                    lf.append(lp)
                    lf.append(t)
                    lf.append(tt)
                    if len(tt)==1:
                        v=list(tt.keys())[0]
                        u=list(tt.values())[0]
                        if u in a[v]:
                            cx=[]
                            for yy in range(len(e[v])):
                                for zz in range(len(e[v][yy])):
                                    if e[v][yy][zz] in a[v] and e[v][yy][zz]!=u:
                                        if str(yy)+str(zz) not in lp:
                                            cx.append(str(yy)+str(zz))
                            if len(cx)!=0:
                                cx=random.choice(cx)
                                e[v][int(cx[0])][int(cx[1])],e[v][int(t[0])][int(t[1])]=e[v][int(t[0])][int(t[1])],e[v][int(cx[0])][int(cx[1])]
                        if u in b[v]:
                            wj=[]
                            for xx in e:
                                for yy in range(len(e[xx])):
                                    for zz in range(len(e[xx][yy])):
                                        if e[xx][yy][zz]==u:
                                            wj.append(str(yy)+str(zz))
                            wj=wj+lp
                            main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
                            ww=[]
                            for y in main:
                                if y not in wj:
                                    ww.append(y)
                            sq={}
                            for xx in e:
                                for yy in range(len(e[xx])):
                                    for zz in range(len(e[xx][yy])):
                                        if str(yy)+str(zz) in ww:
                                            if e[xx][yy][zz] in a[xx]:
                                                sq[xx]=str(yy)+str(zz)
                            if len(sq)!=0:
                                st=random.choice(list(sq))
                                sq=sq[st]
                                e[st][int(t[0])][int(t[1])],e[st][int(sq[0])][int(sq[1])]=e[st][int(sq[0])][int(sq[1])],e[st][int(t[0])][int(t[1])]
        e=final(e, TIME)
        if len(lf)==0:
            foul=1
        if abs(time.time()-TIME)>7:
            e=1
    return e    
def position(d,e,TIME):
    for x in d:
        for y in d[x]:
            for z in d[x][y]:
                if e[x][int(z[0])][int(z[1])]!=y:
                    e=1 
    if abs(time.time()-TIME)>7:
        e=1
    return e
def pr_changer(hh,e,TIME):
    for x in hh:
        fd=[]
        for y in hh[x]:
            if "COMP.PR." in y:
                if y not in fd:
                    fd.append(y)
        if len(fd)==1:
            fd=fd[0]
            for y in range(len(e[x])):
                for z in range(len(e[x][y])):
                    if "COMP.PR." in e[x][y][z]:
                        e[x][y][z]=fd
    if abs(time.time()-TIME)>7:
        e={}
    return e
def assign(e,a,c,TIME):
    for x in e:
        y=count_subjects(e[x])
        z=c[x]
        for u in z:
            if u in y:
                if int(z[u])!=int(y[u]):
                    y=int(z[u])
                    loi=[]
                    lor=[]
                    for xx in range(len(e[x])):
                        for yy in range(len(e[x][xx])):
                            if e[x][xx][yy] in a[x]:
                                loi.append(str(xx)+str(yy))
                            if e[x][xx][yy]==u:
                                lor.append(str(xx)+str(yy))
                    loi=find_pairs(loi)
                    loi=random.sample(loi,int(y/2))
                    lom=[]
                    for xx in loi:
                        for yy in xx:
                            lom.append(yy)
                    loi=[]
                    for xx in range(len(e[x])):
                        for yy in range(len(e[x][xx])):
                            if e[x][xx][yy]=="":
                                loi.append(str(xx)+str(yy))
                    loi=random.sample(loi,len(lom)-len(lor))
                    loi=loi+lor
                    for xx in range(y):
                        yy=loi[xx]
                        zz=lom[xx]
                        e[x][int(yy[0])][int(yy[1])]=e[x][int(zz[0])][int(zz[1])]
                        e[x][int(zz[0])][int(zz[1])]=u
    if abs(time.time()-TIME)>7:
        e=1 
    return e
def out(a):
    trp = 0
    start_time = time.time()
    
    while trp == 0:
        current_time = time.time()
        if (current_time - start_time) > 30:
            trp=2
            e={}
            break
        try:
            TIME = time.time()
            e = output(a, b, c, d, pp, aas, bbs, ccs, ff, TIME)
            e = assign(e, a, c, TIME)
            e = non_movable(e, d, TIME)
            e = pr_changer(hh, e, TIME)
            e = doubler(dx, e, a, TIME)
            e = helop(e, a, b, c, TIME)
            e = position(d, e, TIME)
            e = final(e, TIME)
            if len(e) == len(a):
                trp = 1
                
        except:
            pass
    return e,trp
def poi(d,e,f):
    dq={}
    df=open("INPUT2.csv","r",newline="")
    wr=list(csv.reader(df))
    wd=[]
    wt=[]
    we={}
    wer=[]
    hg=[]
    for x in wr:
        z=[]
        for y in x:
            if y!="":
                z.append(y)
        if z!=[]:
            if z[0][0]=="-":
                z=[]
        if z!=[]:
            wd.append(z)
            if z[0] not in wt:
                wt.append(z[0])
    for x in wt: 
        cx={}
        for y in wd:
            if y[0]==x:
                z=y[2:-1]
                for qq in z:
                    if qq not in hg:
                        hg.append(qq)
                y=y[1]
                cx[y]=z
                
        we[x]=cx
    for x in d:
        sq={}
        for y in d[x]:
            if y in wt:
                sq[y]=d[x][y]
        d[x]=sq
    for x in d:
        if d[x]!={}:
            dq[x]=d[x]
    
    for x in hg:
        am={}
        for xq in dq:
            am[xq]=[]
        tq=[]
        for y in we:
            for z in we[y]:
                if x in we[y][z]:
                    
                    for xx in dq:
                        c=""
                        for zz in dq[xx]:
                            if y in zz:
                                c=zz
                        if c!="":
                            j=c
                            c=dq[xx][c]
                            k=1
                            for yx in c:
                                if yx[0]==z:
                                    k=0
                            if k==0:
                                
                                pm=0
                                for yx in c:
                                    if e[xx][int(yx[0])][int(yx[1])]!=j:
                                        pm=1
                                if pm==0:
                                    for t in c:
                                        if t not in tq:
                                            tq.append(t)
        
        if x in f:
            c=[]
            for y in tq:
                if f[x][int(y[0])][int(y[1])]!="":
                    c.append([f[x][int(y[0])][int(y[1])],y])
            if c!=[]:
                wer.append(c)
    return wer
def ped_no(e,a):
    for x in e:
        for y in range(len(e[x])):
            if e[x][y][9]=="PED":
                mh=[]
                for xx in range(len(e[x])):
                    for yy in range(len(e[x][xx])):
                        if e[x][xx][yy] in a[x]:
                            mh.append(str(xx)+str(yy))
                mh=random.choice(mh)
                e[x][y][9],e[x][int(mh[0])][int(mh[1])]=e[x][int(mh[0])][int(mh[1])],e[x][y][9]
    return e
def fer(e,a,mfq):
    for x in mfq:
         for y in x:
             y=[y[0][0:y[0].index("-")],y[0][y[0].index("-")+1:len(y[0])],y[1]]
             dd=list(a[y[0]].keys())
             cb=[]
             if y[1] in dd:
                 for xx in range(len(e[y[0]])):
                     for yy in range(len(e[y[0]][xx])):
                         if e[y[0]][xx][yy] in dd:
                             if e[y[0]][xx][yy]!=y[1]:
                                 cb.append(str(xx)+str(yy))
                 cb=random.choice(cb)
                 bc=y[2]
                 gh=y[0]
                 e[gh][int(bc[0])][int(bc[1])],e[gh][int(cb[0])][int(cb[1])]=e[gh][int(cb[0])][int(cb[1])],e[gh][int(bc[0])][int(bc[1])]
    return e
def convert_number(num):
    if (num - 3) % 8 == 0:
        return "MON"
    elif (num - 4) % 8 == 0:
        return "TUE"
    elif (num - 5) % 8 == 0:
        return "WED"
    elif (num - 6) % 8 == 0:
        return "THU"
    elif (num - 7) % 8 == 0:
        return "FRI"
    else:
        return "ERROR"
def outer(a,pp,d):  
    trp=0
    crt=0
    while crt==0:
        while trp!=1:
            try:
                e,trp=out(a)
                e=ped_no(e,a)
                e=prt_changer(a,e)
                TIME=time.time()
                e=final(e, TIME)
            except:
                pass
        kcp=[]
        for x in pp:
            mm=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                if mm[yy][zz]=="":
                                    pass
                                else:
                                    if x not in kcp:
                                        kcp.append(x)
        if len(kcp)==0:
            crt=1
    f={}
    for x in pp:
        mm=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
        for y in pp[x]:
            for z in pp[x][y]:
                for yy in range(len(e[y])):
                    for zz in range(len(e[y][yy])):
                        if e[y][yy][zz]==z:
                            if mm[yy][zz]=="":
                                mm[yy][zz]=y+"-"+z
        f[x]=mm
    for wert in range(5):
        mfq=poi(d,e,f)
        e=fer(e,a,mfq)
        e=final(e,time.time())
        f={}
        for x in pp:
            mm=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
            for y in pp[x]:
                for z in pp[x][y]:
                    for yy in range(len(e[y])):
                        for zz in range(len(e[y][yy])):
                            if e[y][yy][zz]==z:
                                if mm[yy][zz]=="":
                                    mm[yy][zz]=y+"-"+z
            f[x]=mm
    mfq=poi(d,e,f)
    return mfq,e,f
    
def otables(e,b,c):
    l={}
    for x in b:
        for y in b[x]:
            if y not in l:
                l[y]=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
    
    for x in c:
        for y in c[x]:
            if y not in l:
                l[y]=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
    l.pop('PED')
    lp=[]
    for y in l:
        for x in e:
            for u in range(len(e[x])):
                for v in range(len(e[x][u])):
                    if y in e[x][u][v]:
                        if l[y][u][v]=="":
                            l[y][u][v]=x
                        else:
                            if y not in lp:
                                lp.append(y)
    return l,lp
def blanker(e):
    lp=0
    for x in e:
        for y in e[x]:
            for z in y:
                if z=="":
                    lp=1
                    break
    return lp
def outest(a,pp,d,b,c,aas,alpha):
    kcit=0
    dw=[]
    tuy=time.time()
    while kcit==0:
        mfq,e,f=outer(a,pp,d)
        e=final(e,time.time())
        dw,lp=otables(e,b,c)
        lm=blanker(e)
        if mfq==[] and lm==0:
            kcit=1
            break
        if time.time()-tuy>5:
            break
    for u in lp:
        lp=[]
        ls={}
        for v in e:
            for w in range(len(e[v])):
                for x in range(len(e[v][w])):
                    if u in e[v][w][x]:
                        if str(w)+str(x) in lp:
                            ls[v]=str(w)+str(x)
                        lp.append(str(w)+str(x))
        main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
        for v in lp:
            if v in main:
                main.remove(v)
        main=find_pairs(main)
        for v in ls:
            ai=ls[v]
            al=""
            try:
                if e[v][int(ai[0])][int(ai[1])-1]==e[v][int(ai[0])][int(ai[1])]:
                    al=ai[0]+str(int(ai[1])-1)
                elif e[v][int(ai[0])][int(ai[1])+1]==e[v][int(ai[0])][int(ai[1])]:
                    al=ai[0]+str(int(ai[1])+1)
            except:
                pass
            lj=[]
            for yy in range(len(e[v])):
                for zz in range(len(e[v][yy])):
                    if e[v][yy][zz] in a[v]:
                        lj.append(str(yy)+str(zz))
            lj=find_pairs(lj)
            lh=[]
            for y in lj:
                if y in main:
                    lh.append(y)
            if len(lh)>0:
                lh=random.choice(lh)
                bi=lh[0]
                bl=lh[1]
                main.remove(lh)
                try:
                    e[v][int(ai[0])][int(ai[1])],e[v][int(bi[0])][int(bi[1])]=e[v][int(bi[0])][int(bi[1])],e[v][int(ai[0])][int(ai[1])]
                    e[v][int(al[0])][int(al[1])],e[v][int(bl[0])][int(bl[1])]=e[v][int(bl[0])][int(bl[1])],e[v][int(al[0])][int(al[1])]
                except:
                    pass
    e=final(e,time.time())
    dw,lp=otables(e,b,c)
    for u in lp:
        lp=[]
        ls={}
        for x in e:
            for y in range(len(e[x])):
                for z in range(len(e[x][y])):
                    if e[x][y][z]==u:
                        if str(y)+str(z) in lp:
                            ls[x]=str(y)+str(z)
                        lp.append(str(y)+str(z))
        for y in ls:
            if ls[y][1]=='7':
                k=[]
                for x in e:
                    if e[x][int(ls[y][0])][int(ls[y][1])]==u:
                        k.append(x)
                k.remove(y)
                kk=[]
                for r in k:
                    if e[r][int(ls[y][0])][int(ls[y][1])-1]==u:
                        kk.append(x)
                if len(k)>0:
                    k=k[0]
                    m=int(ls[y][0])
                    if e[k][m][5] in aas[k]:
                        e[k][m][5],e[k][m][7]=e[k][m][7],e[k][m][5]
            elif ls[y][1]=='6':
                k=[]
                for x in e:
                    if e[x][int(ls[y][0])][int(ls[y][1])]==u:
                        k.append(x)
                k.remove(y)
                kk=[]
                for r in k:
                    if e[r][int(ls[y][0])][int(ls[y][1])+1]==u:
                        kk.append(x)
                if len(k)>0:
                    k=k[0]
                    m=int(ls[y][0])
                    if e[k][m][8] in aas[k]:
                        e[k][m][6],e[k][m][8]=e[k][m][6],e[k][m][8]
                    
    e=final(e,time.time())
    dw,lp=otables(e,b,c)
    for x in lp:
        main=['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49']
        for u in e:
            for v in range(len(e[u])):
                for w in range(len(e[u][v])):
                    if x in e[u][v][w]:
                        if str(v)+str(w) in main:
                            main.remove(str(v)+str(w))
        main=find_pairs(main)
        lp=[]
        ls={}
        for u in e:
            for y in range(len(e[u])):
                for z in range(len(e[u][y])):
                    if x in e[u][y][z]:
                        if str(y)+str(z) in lp:
                            ls[u]=str(y)+str(z)
                        lp.append(str(y)+str(z))
        for y in ls:
            lt=ls[y]
            lj=""
            for u in e:
                if x in e[u][int(ls[y][0])][int(ls[y][1])]:
                    lj=u
                    break
            lm=[]
            for yy in range(len(e[lj])):
                for zz in range(len(e[lj][yy])):
                    if e[lj][yy][zz] in a[lj]:
                        lm.append(str(yy)+str(zz))
            lm=find_pairs(lm)
            lk=[]
            for y in lm:
                if y in main:
                    lk.append(y)
            if len(lk)>0:
                lk=random.choice(lk)
                main.remove(lk)
                ln=[]
                for yy in range(len(e[lj])):
                    for zz in range(len(e[lj][yy])):
                        if x in e[lj][yy][zz]:
                            ln.append(str(yy)+str(zz))
                ln=find_pairs(ln)
                lg=[]
                for y in ln:
                    if lt in y:
                        lg=y
                try:
                    ai=lk[0]
                    al=lk[1]
                    bi=lg[0]
                    bl=lg[1]
                    e[lj][int(ai[0])][int(ai[1])],e[lj][int(bi[0])][int(bi[1])]=e[lj][int(bi[0])][int(bi[1])],e[lj][int(ai[0])][int(ai[1])]
                    e[lj][int(al[0])][int(al[1])],e[lj][int(bl[0])][int(bl[1])]=e[lj][int(bl[0])][int(bl[1])],e[lj][int(al[0])][int(al[1])]
                except:
                    pass
    e=final(e,time.time())
    dw,lp=otables(e,b,c)
    for u in lp:
        jj=['05','08','15','18','25','28','35','38','45','48']
        gc=[]
        sc={}
        for v in e:
            for w in range(len(e[v])):
                for x in range(len(e[v][w])):
                    if u in e[v][w][x]:
                        if str(w)+str(x) in jj:
                            jj.remove(str(w)+str(x))
                        if str(w)+str(x) in gc:
                            if v in sc:
                                sc[v]=sc[v]+[str(w)+str(x)]
                            else:
                                sc[v]=[str(w)+str(x)]
                        gc.append(str(w)+str(x))

        jm=[['05','08'],['15','18'],['25','28'],['35','38'],['45','48']]
        kh=[]
        for y in jm:
            if y[0] in jj and y[1] in jj:
                kh.append(y)
        if len(kh)>0:
            if len(sc)>0:
                for y in kh:
                    jg=int(y[0][0])
                    ui=""
                    for x in e:
                        if u in e[x][jg][6] and u in e[x][jg][7] and e[x][jg][5] in a[x]:
                            ui=x
                    e[ui][jg][5],e[ui][jg][7]=e[ui][jg][7],e[ui][jg][5]
                    er=random.choice(list(sc))
                    em=sc[er]
                    try:
                        if er in sc:
                            sc.pop(er)
                        if e[er][jg][7] in a[x] and e[er][jg][8] in a[x]:
                            ek=em[0]
                            et=em[1]
                            e[er][jg][7],e[er][int(ek[0])][int(ek[1])]=e[er][int(ek[0])][int(ek[1])],e[er][jg][7]
                            e[er][jg][8],e[er][int(et[0])][int(et[1])]=e[er][int(et[0])][int(et[1])],e[er][jg][8]
                    except:
                        try:
                            mn=[]
                            for xx in range(len(e[er])):
                                for yy in range(len(e[er][xx])):
                                    if u in e[er][xx][yy]:
                                        mn.append(str(xx)+str(yy))
                            em=mn[2]
                            if er in sc:
                                sc.pop(er)
                            if e[er][jg][7] in a[x] and e[er][jg][8] in a[x]:
                                ek=em[0]
                                et=em[1]
                                e[er][jg][7],e[er][int(ek[0])][int(ek[1])]=e[er][int(ek[0])][int(ek[1])],e[er][jg][7]
                                e[er][jg][8],e[er][int(et[0])][int(et[1])]=e[er][int(et[0])][int(et[1])],e[er][jg][8]
                        except:
                            pass
    if time.time()-alpha>15:
        e=1                
    e=final(e,time.time())
    dw,lp=otables(e,b,c)
    
    return e,dw,lp  
def printer(a,pp,d,b,c,aas,alpha):
    dqw=0
    while dqw==0:
        e,dw,lp=outest(a,pp,d,b,c,aas,alpha)
        bc=blanker(e)
        if bc==0:
            dqw=1
            break
        if time.time()-alpha>15:
            e=1
    return e,dw
df=open("INPUT0.csv","r",newline="")
wr=csv.reader(df)
L=[]
for x in wr:
    while "" in x:
        x.remove("")
    if len(x)>0:
        if x[0][0]=="-":
            pass
        elif x[0][0]=="" or x[0]=='TOTAL':
            pass
        elif x[0]=='BREAK' :
            L.append(x[0:2])
        elif str(x[0][0]).upper()=='V' or str(x[0][0]).upper()=='I' or str(x[0][0]).upper()=='X':
            L.append(x[0])
        else:
            L.append(x)
df.close()
couy=1
cooz=1
H=[]
M=[]
for x in range(len(L)):
    ctp=(5*couy)-3
    ctt=(5*cooz)-2
    if x==ctp:
        lk=[]
        for y in L[x]:
            lk.append(y)
        H.append(lk)
        couy+=1
    elif x==ctt:
        lm=convert_string_list(L[x])
        H.append(lm)
        cooz+=1
    else:
        H.append(L[x])
    if L[x]==str(L[x]):
        M.append(L[x])
DC={}
for x in M:
    n=H.index(x)
    p=list(zip(H[n+1],H[n+2]))
    p= tuples_to_dict(p)
    c=0
    for y in p:
        if "{U}" in y:
            p[y]=H[n+3][c]
            c+=1
    DC[H[n]]=p
a={}
b={}
c={}
d={}
dx={}
for x in DC:
    p=DC[x]
    a1={}
    b1={}
    c1={}
    d1={}
    dx1={}
    for y in p:
        if "{" not in y:
            a1[y]=p[y]
        elif "{U}" in y:
            nk=y.index("{")
            f=y[0:nk]
            d1[f]=p[y]
        elif "{S}" in y:
            if "{D}" in y:
                nk=y.index("{")
                f=y[0:nk]
                if "COMP.PR.1" in f:
                    f="COMP.PR.1"
                if "COMP.PR.2" in f:
                    f="COMP.PR.2"
                c1[f]=p[y]
            else:
                nk=y.index("{")
                f=y[0:nk]
                b1[f]=p[y]
        elif "{D}" in y and "{S}" not in y:
            nk=y.index("{")
            f=y[0:nk]
            dx1[f]=p[y]
    a[x]=a1
    b[x]=b1
    c[x]=c1
    d[x]=d1
    if dx1!={}:
        dx[x]=dx1
for x in M:
    y=L.index(x)
    f=L[y+4][1]
    d[x].update({"BREAK":["0"+f,"1"+f,"2"+f,"3"+f,"4"+f]})
aas=a.copy()
bbs=b.copy()
ccs=c.copy()
dds=d.copy()
zz={}
for x in aas:
    zz[x]=aas[x] |bbs[x] | ccs[x] | dds[x]
vv=[]
for x in zz:
    for v in zz[x].keys():
        if v not in vv:
            vv.append(v)
ed={}
for x in d:
    D={}
    for y in d[x]:
        D[y]=len(d[x][y])
    ed[x]=D
ff={}
for x in a:
    ff[x]={}
    ff[x].update(a[x])
    ff[x].update(b[x])
    ff[x].update(c[x])
    ff[x].update(ed[x])
ds=[]
dm=[]
df=open("INPUT1.csv","r",newline="")
wr=csv.reader(df)
for x in wr:
    while "" in x:
        x.remove("")
    if x!=[]:
        if x[0][0]!="-":
            if len(x)==2:
                x=[{x[0]:x[1]}]
            elif len(x)==3:
                x=[{x[0]:x[1]},{x[0]:x[2]}]
            elif len(x)==4:
                x=[{x[0]:x[1]},{x[0]:x[2]},{x[0]:x[3]}]
            elif len(x)==5:
                x=[{x[0]:x[1]},{x[0]:x[2]},{x[0]:x[3]},{x[0]:x[4]}]
            else:
                x=x[0]
                dm.append(x)
            ds.append(x)
ii={}
hh={}
for x in dm:
    if x[0]=="I" or x[0]=="V" or x[0]=="X":
        n=ds.index(x)
        j=1
        D={}
        DS={}
        while ds[n+j]==list(ds[n+j]):
            m=ds[n+j]
            for y in m:
                h=list(y)[0]
                if h in a[x]:
                    D.update(y)
                DS.update(y)
            if n+j==len(ds)-1:
                break
            j+=1
        if x[0]=="I" or x[0]=="V" or x[0]=="X":
            ii[x]=D
            hh[x]=DS
ln=[]
for x in ds:
    if x==list(x):
        for y in x:
            for z in y:
                if y[z] not in ln:
                    ln.append(y[z])

for x in ln:
    count=0
    lj=[]
    for y in hh:
        for z in hh[y]:
            if hh[y][z]==x:
                lj.append(z)
                if z in ff[y]:
                    kj=int(ff[y][z])
                    count+=kj
jj={}
for x in ln:
    dw={}
    for y in hh:
        for z in hh[y]:
            if hh[y][z]==x:
                if z in ff[y]:
                    if z not in b[y] and z not in d[y]:
                        if y not in dw:
                            dw[y]={z:ff[y][z]}
                        else:
                            dw[y].update({z:ff[y][z]})
                    
    jj[x]=dw
df=open("INPUT1.csv","r",newline="")
wr=list(csv.reader(df))
L=[]
for x in wr:
    o=[]
    for y in x:
        if y!="":
            if len(y)!=0:
                if y[0]!="-" and y[1]!="-" and y[-1]!="-":
                    o.append(y)
    if o!=[]:
        L.append(o)
df.close()
Y={}
for x in L:
    if len(x)==1:
        if x[0][0]=="V" or x[0][0]=="I" or x[0][0]=="X":
            Y[x[0]]=[]
Z=list(Y.keys())
for x in L:
    V=-1
    if x[0] in Z:
        V=L.index(x)
    if V!=-1:
        LY=[]
        for y in range(V+1,len(L)):
            ds={L[y][0]:L[y][1:len(L[y])]}
            LY.append(ds)
            if y<len(L)-1:
                if L[y+1][0] in Z:
                    break
        Y[x[0]]=LY
T={}
for x in Y:  
    for y in Y[x]:
        y=list(y.values())[0]
        for z in y:
            if z not in T:
                T[z]={}
pp={}
for x in T:
    ds={}
    for y in Y:
        dm=[]
        for z in Y[y]:
            q=list(z.keys())[0]
            z=list(z.values())[0]
            if x in z:
                dm.append(q)
        if dm!=[]:
            ds[y]=dm
    pp[x]=ds
    
erteiurtyiery=0 
while erteiurtyiery==0:
    e=[]
    while len(e)!=len(aas):
        try:
            alpha=time.time()
            e,dw=printer(a,pp,d,b,c,aas,alpha)
        except:
            break
    if len(e)==len(aas):
        f={}
        for x in pp:
            mm=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
            for y in pp[x]:
                for z in pp[x][y]:
                    try:
                        for yy in range(len(e[y])):
                            for zz in range(len(e[y][yy])):
                                if e[y][yy][zz]==z:
                                    if mm[yy][zz]=="":
                                        mm[yy][zz]=y+"-"+z
                    except:pass
            f[x]=mm
        dq={}
        df=open("INPUT2.csv","r",newline="")
        wr=list(csv.reader(df))
        wd=[]
        wt=[]
        we={}
        hg=[]
        for x in wr:
            z=[]
            for y in x:
                if y!="":
                    z.append(y)
            if z!=[]:
                if z[0][0]=="-":
                    z=[]
            if z!=[]:
                wd.append(z)
                if z[0] not in wt:
                    wt.append(z[0])
        for x in wt: 
            cx={}
            for y in wd:
                if y[0]==x:
                    z=y[2:-1]
                    for qq in z:
                        if qq not in hg:
                            hg.append(qq)
                    y=y[1]
                    cx[y]=z
                    
            we[x]=cx
        for x in d:
            sq={}
            for y in d[x]:
                if y in wt:
                    sq[y]=d[x][y]
            d[x]=sq
        for x in d:
            if d[x]!={}:
                dq[x]=d[x]
        
        for x in hg:
            am={}
            for xq in dq:
                am[xq]=[]
            tq=[]
            for y in we:
                for z in we[y]:
                    if x in we[y][z]:
                        
                        for xx in dq:
                            c=""
                            for zz in dq[xx]:
                                if y in zz:
                                    c=zz
                            if c!="":
                                j=c
                                c=dq[xx][c]
                                k=1
                                for yx in c:
                                    if yx[0]==z:
                                        k=0
                                if k==0:
                                    try:
                                        pm=0
                                        for yx in c:
                                            if e[xx][int(yx[0])][int(yx[1])]!=j:
                                                pm=1
                                        if pm==0:
                                            for t in c:
                                                if t not in tq:
                                                    tq.append(t)
                                            am[xx]=am[xx]+[j]
                                    except:
                                        pass
            am=list(am.values())
            er=[]
            for y in am:
                if y not in er:
                    if y!=[]:
                        er.append(y)
            er=er[0][0]
            if x in f:
                for z in tq:
                    em=""
                    for xx in e:
                        if e[xx][int(z[0])][int(z[1])]==er:
                            em=xx[0:xx.index(" ")]
                            break
                            
                    f[x][int(z[0])][int(z[1])]=em+"-"+er
            else:
                 f[x]=[["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""],["","","","","","","","","",""]]
                 for z in tq:
                     em=""
                     for xx in e:
                         if e[xx][int(z[0])][int(z[1])]==er:
                             em=xx[0:xx.index(" ")]
                             break
                             
                     f[x][int(z[0])][int(z[1])]=em+"-"+er
        aa=[]
        bb=[] 
        cc=[]      
        for x in e:
            aa.append([x,'','','','','','','','','CT'+"-"+hh[x]['CT'],'CO:CT'+"-"+hh[x]['CO-CT']])
            aa.append(['X','-0-','-1-','-2-','-3-','-4-','-5-','-6-','-7-','-8-','-9-'])
            for y in e[x]:
                SS=convert_number(len(aa)+1)
                y=[SS]+y
                aa.append(y)
            aa.append([])
        for x in f:
            bb.append([x])
            bb.append(['X','-0-','-1-','-2-','-3-','-4-','-5-','-6-','-7-','-8-','-9-'])
            for y in f[x]:
                SS=convert_number(len(bb)+1)
                y=[SS]+y
                bb.append(y)
            bb.append([])
        for x in dw:
            cc.append([x])
            cc.append(['X','-0-','-1-','-2-','-3-','-4-','-5-','-6-','-7-','-8-','-9-'])
            for y in dw[x]:
                SS=convert_number(len(cc)+1)
                y=[SS]+y
                cc.append(y)
            cc.append([])
        a=open("CLASS_ROUTINE.csv","w",newline="")
        b=csv.writer(a)
        b.writerows(aa)
        a.close()
        a=open("TEACHER_ROUTINE.csv","w",newline="")
        b=csv.writer(a)
        b.writerows(bb)
        a.close()
        a=open("SPECIAL_ROUTINE.csv","w",newline="")
        b=csv.writer(a)
        b.writerows(cc)
        a.close()
        class_csv_file = 'CLASS_ROUTINE.csv'
        teacher_csv_file = 'TEACHER_ROUTINE.csv'
        special_csv_file = 'SPECIAL_ROUTINE.csv'
        class_excel_file = 'OUTPUT\\CLASS_ROUTINE.xlsx'
        teacher_excel_file = 'OUTPUT\\TEACHER_ROUTINE.xlsx'
        special_excel_file = 'OUTPUT\\SPECIAL_ROUTINE.xlsx'
        def read_csv(file_path):
            data = []
            with open(file_path, mode='r') as file:
                reader = csv.reader(file)
                for row in reader:
                    data.append(row)
            return data
        def write_to_excel(data, excel_file):
            wb = Workbook()
            ws = wb.active
            for row in data:
                ws.append(row)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = 17
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 22.5
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment
            total_rows = ws.max_row
            row_ranges = [(2, 7), (10, 15), (18, 23), (26, 31)]
            section_height = 8
            for start_row, end_row in row_ranges:
                while end_row <= total_rows:
                    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=11):
                        for cell in row:
                            cell.border = thin_border
                    start_row += section_height
                    end_row += section_height
            wb.save(excel_file)
        for csv_file, excel_file in [
            (class_csv_file, class_excel_file),
            (teacher_csv_file, teacher_excel_file),
            (special_csv_file, special_excel_file),
        ]:
            data = read_csv(csv_file)
            write_to_excel(data, excel_file)
        for csv_file in [class_csv_file, teacher_csv_file, special_csv_file]:
            if os.path.exists(csv_file):
                os.remove(csv_file)
            else:
                pass
        tm=time.time()-tmerytire
        tm="Output saved in "+str(round(tm,3))+"s"
        print(tm)
        break
    else:
        erteiurtyiery=1
        print("Please Run the code again")